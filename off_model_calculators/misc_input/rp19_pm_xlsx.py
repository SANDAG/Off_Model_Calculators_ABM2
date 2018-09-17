from datetime import datetime
import pandas as pd
import openpyxl  # necessary for pandas ExcelWriter
import pyodbc
import sys
import time

# Start Script Timer
start = time.time()


def scenario_info(scenario_id, row_start):
    row_start -= 1
    scenario = pd.read_sql_query(
        sql=("SELECT RTRIM([name]) AS [name], [year] "
             "FROM [dimension].[scenario] WHERE [scenario_id] = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = scenario.at[0, "name"]  # row & column numbers are 0-based
    sheetname.cell(row=row_start+1, column=col).value = scenario_id  # row & column numbers are 0-based


def tot_hh(scenario_id, row_start):
    row_start -= 1
    scenario = pd.read_sql_query(
        sql=("SELECT Count(household_id) AS total_hh "
             "FROM [dimension].[household] WHERE [scenario_id] = ? and [unittype] = 'Non-Group Quarters'"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = scenario.at[0, "total_hh"]

def tot_pop(scenario_id, row_start):
    row_start -= 1
    df_pop = pd.read_sql_query(
        sql=("SELECT SUM([person].[weight_person]) AS [total_pop]"
		     ",SUM(CASE WHEN [household].[poverty] <= 2 THEN [person].[weight_person] ELSE 0 END) AS [pop_low_income]"
		     ",SUM(CASE	WHEN [person].[race] IN ('Some Other Race Alone', " \
												'Asian Alone', \
												'Black or African American Alone', \
												'Two or More Major Race Groups', \
												'Native Hawaiian and Other Pacific Islander Alone', \
												'American Indian and Alaska Native Tribes specified; or American Indian or Alaska Native, not specified and no other races')"
							"OR [person].[hispanic] = 'Hispanic' THEN [person].[weight_person] "
							"ELSE 0 END) AS [pop_minority] "
		    ",SUM(CASE WHEN [person].[age] >= 75 THEN [person].[weight_person] ELSE 0 END) AS [pop_senior] "
	    "FROM [dimension].[person]"
	    "INNER JOIN  [dimension].[household]"
	    "ON [person].[scenario_id] = [household].[scenario_id] "
		"AND [person].[household_id] = [household].[household_id] "
        "FROM [dimension].[person] WHERE [scenario_id] = ?"),
        con=sql_con,
        params=[scenario_id]
    )
	i=0
	for j in xrange(4):
        i += 1
        sheetname.cell(row=row_start + i, column=col).value = \
            df_pop[0][j]]
#    sheetname.cell(row=row_start, column=col).value = scenario.at[0, "total_pop"]
#	sheetname.cell(row=row_start+1, column=col).value = scenario.at[0, "pop_low_income"]
#	sheetname.cell(row=row_start+2, column=col).value = scenario.at[0, "pop_minority"]
#	sheetname.cell(row=row_start+3, column=col).value = scenario.at[0, "pop_senior"]

def tot_emp(scenario_id, row_start):
    row_start -= 1
    scenario = pd.read_sql_query(
	    sql=("SELECT  SUM(EMP_TOTAL) emp_total,SUM(COLLEGEENROLL+OTHERCOLLEGEENROLL) coll_enroll "
             "FROM [fact].[mgra_based_input] WHERE [scenario_id] = ?"),
        con=sql_con,
        params=[scenario_id]
		)
    sheetname.cell(row=row_start, column=col).value = scenario.at[0, "emp_total"]
	sheetname.cell(row=row_start+1, column=col).value = scenario.at[0, "coll_enroll"]

def tot_trips(scenario_id, row_start):
    row_start -= 1
    scenario = pd.read_sql_query(
		sql=("SELECT sum(weight_person_trip) simulated_trips"
             "FROM [fact].[person_trip] WHERE [scenario_id] = ?"
		con=sql_con,
        params=[scenario_id]
		)
    sheetname.cell(row=row_start, column=col).value = scenario.at[0, "simulated_trips"]
	
def pm_1a(scenario_id, row_start):
    row_start -= 1
    df_veh_delay = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_1a] @scenario_id = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = df_veh_delay.iloc[0][1]  # row & column numbers are 0-based


def pm_2a(scenario_id, row_start):
    row_start -= 1
    i = 0
    for index in xrange(len(uats)):
        i += 1
        df_ms_region_alltrips = pd.read_sql_query(
            sql=("EXECUTE [rtp_2019].[sp_pm_2a] @scenario_id = ?, @uats = ?, @work = ?"),
            con=sql_con,
            params=[scenario_id, uats[index], work[index]]
        )
        if work[index] == 0: # all trips
            sheetname.cell(row=row_start + 7 * (i - 1), column=col).value = \
                df_ms_region_alltrips['pct_person_trips'][mode_ix[1]] + \
                df_ms_region_alltrips['pct_person_trips'][mode_ix[2]] + \
                df_ms_region_alltrips['pct_person_trips'][mode_ix[3]] + \
                df_ms_region_alltrips['pct_person_trips'][mode_ix[4]]  # add HOV+Transit+Bike+Walk
            for j in xrange(len(mode_ix)):
                sheetname.cell(row=row_start + 1 + j + 7 * (i - 1), column=col).value = \
                    df_ms_region_alltrips['pct_person_trips'][mode_ix[j]]
        else: # work trips
            sheetname.cell(row=row_start + 7 * (i - 1), column=col).value = \
                df_ms_region_alltrips['pct_person_trips'][mode_work[1]] + \
                df_ms_region_alltrips['pct_person_trips'][mode_work[2]] + \
                df_ms_region_alltrips['pct_person_trips'][mode_work[3]] + \
                df_ms_region_alltrips['pct_person_trips'][mode_work[4]]  # add HOV+Transit+Bike+Walk
            for j in xrange(len(mode_work)):
                sheetname.cell(row=row_start + 1 + j + 7 * (i - 1), column=col).value = \
                    df_ms_region_alltrips['pct_person_trips'][mode_work[j]]

def pm_2b(scenario_id, row_start):
    row_start -= 1
    df_veh_delay = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_2b] @scenario_id = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = df_veh_delay.iloc[0][2]  # row & column numbers are 0-based
    sheetname.cell(row=row_start + 1, column=col).value = df_veh_delay.iloc[0][1]  # row & column numbers are 0-based

def pm_6a(scenario_id, row_start):
    row_start -= 1
    df_ms_region_alltrips = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_6a] @scenario_id = ?, @senior = 0, @minority = 0, @low_income = 0"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = df_ms_region_alltrips['physical_activity_per_capita'][0]
    write_sequence = [0, 1, 0, 1, 1, 0]
    i = 0
    for index in xrange(len(senior)):
        df_ms_region_alltrips = pd.read_sql_query(
            sql=("EXECUTE [rtp_2019].[sp_pm_6a] @scenario_id = ?, @senior = ?, @minority = ?, @low_income = ?"),
            con=sql_con,
            params=[scenario_id, senior[index], minority[index], lowinc[index]]
        )
        for j in xrange(2):
            i += 1
            sheetname.cell(row=row_start + i, column=col).value = \
                df_ms_region_alltrips['physical_activity_per_capita'][write_sequence[i-1]]

def pm_a(scenario_id, row_start):
    row_start -= 1   
    df_time_work_trips = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_A] @scenario_id = ?,@senior = 0, @minority = 0, @low_income = 0"),
        con=sql_con,
        params=[scenario_id]
    )
    mode_work_time = [5, 1, 2, 3, 0, 4]  # travel time work trips: total, sov, hov, transit, bike, walk
    for j in xrange(len(mode_work_time)):
        row_increment = row_start + j
        sheetname.cell(row=row_increment, column=col).value = \
            df_time_work_trips['avg_time_trip'][mode_work_time[j]]

    
    write_sequence = [5, 1, 2, 3, 0, 4, 11, 7, 8, 9, 6, 10] # for low income and non-low, minority and non-minority
    for index in xrange(len(senior)-1):  # only run income and minority, no senior
        df_time_work_trips = pd.read_sql_query(
            sql=("EXECUTE [rtp_2019].[sp_pm_A] @scenario_id = ?,@senior = ?, @minority = ?, @low_income = ?"),
            con=sql_con,
            params=[scenario_id, senior[index], minority[index], lowinc[index]]
        )
        
        for j in xrange(len(write_sequence)):
            row_increment += 1 
            sheetname.cell(row=row_increment, column=col).value = \
                df_time_work_trips['avg_time_trip'][write_sequence[j]]	
            

def pm_b(scenario_id, row_start):
    row_start -= 1
    df_tribal_time = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_B] @scenario_id = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = df_tribal_time.iloc[0][1]  # row & column numbers are 0-based


def pm_c(scenario_id, row_start):
    row_start -= 1
    scenario = pd.read_sql_query(
        sql=("SELECT RTRIM([name]) AS [name], [year] "
             "FROM [dimension].[scenario] WHERE [scenario_id] = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    df_time = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_C] @scenario_id = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    write_sequence_preOME = [3, 1, 0, 2]  # total, San Ysidro, Otay Mesa, Tecate
    write_sequence_postOME = [4, 2, 0, 1, 3]  # total, San Ysidro, Otay Mesa, OME, Tecate
    i=0
    if scenario.at[0, "year"] < 2035:  # prior to 2035, OME is not available
        for index in xrange(len(write_sequence_preOME)):		
            sheetname.cell(row=row_start+i, column=col).value = \
                df_time.iloc[write_sequence_preOME[index]][2]  # row & column numbers are 0-based
            i += 1
            if i == 3:  # skip OME row as data doesnot exist
                i += 1
    else:
        for index in xrange(len(write_sequence_postOME)):		
            sheetname.cell(row=row_start+i, column=col).value = \
			df_time.iloc[write_sequence_postOME[index]][2]  # ro
            i +=1
			
def pm_d(scenario_id, row_start):
    row_start -= 1
    df_time = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_D] @scenario_id = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = df_time.iloc[0][1]  # row & column numbers are 0-based

def pm_e(scenario_id, row_start):
    row_start -= 1
    df_time = pd.read_sql_query(
        sql=("EXECUTE [rtp_2019].[sp_pm_E] @scenario_id = ?"),
        con=sql_con,
        params=[scenario_id]
    )
    sheetname.cell(row=row_start, column=col).value = df_time.iloc[0][1]  # row & column numbers are 0-based

def pm_f(scenario_id, row_start):
    row_start -= 1
    df_pct_cost = pd.read_sql_query(
			sql=("EXECUTE [rtp_2019].[sp_pm_F] @scenario_id = ?, @senior = 0, @minority = 0, @low_income = 0"),
			con=sql_con,
			params=[scenario_id]
	)
    sheetname.cell(row=row_start, column=col).value = df_pct_cost['pct_income_transportation_cost'][0]
    write_sequence = [0, 1, 0, 1, 1, 0]
    i = 0
    for index in xrange(len(senior)):
        df_pct_cost = pd.read_sql_query(
			sql=("EXECUTE [rtp_2019].[sp_pm_F] @scenario_id = ?, @senior = ?, @minority = ?, @low_income = ?"),
			con=sql_con,
			params=[scenario_id, senior[index], minority[index], lowinc[index]]
		)		
        for j in xrange(2):
            i += 1
            sheetname.cell(row=row_start + i, column=col).value = \
				df_pct_cost['pct_income_transportation_cost'][write_sequence[i-1]] 

def pm_h(scenario_id, row_start):
    row_start -= 1
    df_pct_pop_physical = pd.read_sql_query(
			sql=("EXECUTE [rtp_2019].[sp_pm_H] @scenario_id = ?, @senior = 0, @minority = 0, @low_income = 0"),
			con=sql_con,
			params=[scenario_id]
	)
    sheetname.cell(row=row_start, column=col).value = \
	    df_pct_pop_physical['pct_physical_activity_population'][0]
    write_sequence = [0, 1, 0, 1, 1, 0]
    i = 0
    for index in xrange(len(senior)):
        df_pct_pop_physical = pd.read_sql_query(
			sql=("EXECUTE [rtp_2019].[sp_pm_H] @scenario_id = ?, @senior = ?, @minority = ?, @low_income = ?"),
			con=sql_con,
			params=[scenario_id, senior[index], minority[index], lowinc[index]]
		)
        for j in xrange(2):
            i += 1
            sheetname.cell(row=row_start + i, column=col).value = \
				df_pct_pop_physical['pct_physical_activity_population'][write_sequence[i-1]] 
"""    	
def pm_vmt(scenario_id, row_start):
    row_start -= 1
    df_vmt = pd.read_sql_query(
			sql=("EXECUTE [rtp_2019].[sp_pm_3ab_vmt] @scenario_id = ?"),
			con=sql_con,
			params=[scenario_id]
	)
"""	
		
filename = 'Scenario SDF Performance Measures_template'
path = 'T:/RTP/2019RP/rp19_scen/analysis/'
srcfile = openpyxl.load_workbook(path + filename + '.xlsx')

sheetname = srcfile.get_sheet_by_name('PM_ABM14.0.0')

# set sql server connection string
# noinspection PyArgumentList
sql_con = pyodbc.connect(driver='{SQL Server}',
                         server='sql2014a8',
                         database='abm_2',
                         trusted_connection='yes')

col = 1
uats = [0, 0, 1, 1]  # binary - 1 is UATS, 0 is all areas
work = [0, 1, 0, 1]  # binary - 1 is work, 0 is all trips
senior = [0, 0, 1]  # binary - 1 is senior, 0 is non-senior
minority = [0, 1, 0]  # binary - 1 is minority, 0 is non-minority
lowinc = [1, 0, 0]  # binary - 1 is low income, 0 is non-low income
mode_ix = [1, 3, 4, 0, 5, 2]  # all trips: sov, hov, transit, bike, walk, other
mode_work = [1, 2, 3, 0, 4]  # work trips: sov, hov, transit, bike, walk
for id_argv in sys.argv[1:]:
    scenario_id = int(id_argv)
    col += 1
    # get scenario information for the given scenario_id
    print "Scenario Info: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    scenario_info(scenario_id, 1)  # row is excel row number

    # get total households
    print "Total HH: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    tot_hh(scenario_id, 3)  # row is excel row number
    
    # get total pop
    print "Total Pop: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    tot_pop(scenario_id, 4)  # row is excel row number

	# get total EMP
    print "Total Pop: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    tot_emp(scenario_id, 8)  # row is excel row number
	
	# get simulated trips
    print "Total Pop: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    tot_trips(scenario_id, 13)  # row is excel row number
	
	# get VMT
    print "PM2b ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_2b(scenario_id, 10)  # row is excel row number
	
    # get query results for the given scenario_id
    print "PM1a ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_1a(scenario_id, 19)  # row is excel row number

    print "PM2a ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_2a(scenario_id, 20)  # row is excel row number

    # print "PM3a ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    # print "PM3b ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)

    print "PM6a ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_6a(scenario_id, 67)  # row is excel row number

    # print "PM7a ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    # print "PM7b ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
 
    print "PM_A ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_a(scenario_id, 187)  # row is excel row number

	
    print "PM_B ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_b(scenario_id, 217)  # row is excel row number

    print "PM_C ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_c(scenario_id, 218)  # row is excel row number
	
    print "PM_D ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_d(scenario_id, 223)  # row is excel row number

    print "PM_E ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_e(scenario_id, 224)  # row is excel row number
    
    print "PM_F ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_f(scenario_id, 225)  # row is excel row number
	
    print "PM_H ID: " + str(scenario_id) + "  Elapsed time: %5.2f mins" % ((time.time() - start) / 60.0)
    pm_h(scenario_id, 240)  # row is excel row number
 
fileout = path + filename[:-9].strip() + str(datetime.now().strftime("%Y-%m-%d") + '.xlsx')
srcfile.save(fileout)
print "Finished in %5.2f mins" % ((time.time() - start)/60.0)
